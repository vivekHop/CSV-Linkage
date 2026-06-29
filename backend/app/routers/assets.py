from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from sqlalchemy.orm import Session
from typing import List, Dict, Any
from datetime import datetime
from app.database import get_db
from app.schemas import AssetResponse, AssetUpdate, VersionHistoryResponse, AssetCreate, WorkspaceSync, ImportDraftCreate, ImportDraftResponse
from app.models import Asset, ColumnModel, RelationshipModel
from app.repositories import AssetRepository, ColumnRepository, RelationshipRepository, ImportDraftRepository
from app.profiler import profile_file

from app.websockets import manager

router = APIRouter(prefix="/assets", tags=["Assets"])

@router.post("/upload", response_model=List[AssetResponse], status_code=status.HTTP_201_CREATED)
async def upload_spreadsheet_files(files: List[UploadFile] = File(...), db: Session = Depends(get_db)):
    """
    Uploads and profiles one or more spreadsheets (CSV, Excel, TSV, ODS).
    If a file contains multiple sheets (like Excel), profiles and creates a separate table for each sheet.
    The raw data is processed in-memory and discarded; only the metadata is stored.
    """
    asset_repo = AssetRepository(db)
    column_repo = ColumnRepository(db)
    created_assets = []

    allowed_extensions = (".xlsx", ".xls", ".xlsm")

    for file in files:
        if not file.filename.lower().endswith(allowed_extensions):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"File '{file.filename}' is not a supported Excel workbook format (.xlsx, .xls, .xlsm)."
            )
        
        try:
            # Read file bytes in-memory
            file_bytes = await file.read()
            
            # Profile the spreadsheet file (returns a list of sheet profiles)
            profiled_sheets = profile_file(file_bytes, file.filename)
            
            for asset_data, columns_data in profiled_sheets:
                # Save Asset
                asset = asset_repo.create(
                    name=asset_data["name"],
                    asset_type=asset_data["asset_type"],
                    row_count=asset_data["row_count"],
                    column_count=asset_data["column_count"],
                    file_size=asset_data["file_size"],
                    description=asset_data["description"],
                    owner=asset_data["owner"],
                    notes=asset_data["notes"],
                    tags=asset_data["tags"],
                    custom_attributes=asset_data["custom_attributes"]
                )
                
                # Save Columns
                for col_data in columns_data:
                    column_repo.create(asset.id, col_data)
                    
                # Fetch complete asset with columns populated
                db_asset = asset_repo.get_by_id(asset.id)
                created_assets.append(db_asset)
            
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Failed to profile and import '{file.filename}': {str(e)}"
            )
            
    # Broadcast upload events
    for asset in created_assets:
        await manager.broadcast({"event_type": "asset_created", "data": {"id": asset.id}})

    return created_assets

@router.post("", response_model=AssetResponse, status_code=status.HTTP_201_CREATED)
async def create_asset(asset: AssetCreate, db: Session = Depends(get_db)):
    """
    Creates a new asset and its columns directly from a metadata payload.
    Used for canvas duplicating, copy-pasting, and grouping.
    """
    asset_repo = AssetRepository(db)
    column_repo = ColumnRepository(db)
    
    created_asset = asset_repo.create(
        name=asset.name,
        asset_type=asset.asset_type,
        row_count=asset.row_count,
        column_count=asset.column_count,
        file_size=asset.file_size,
        description=asset.description or "",
        owner=asset.owner or "",
        notes=asset.notes or "",
        tags=asset.tags or [],
        custom_attributes=asset.custom_attributes or {}
    )
    
    for col in asset.columns:
        column_repo.create(created_asset.id, {
            "name": col.name,
            "datatype": col.datatype,
            "nullable_percentage": col.nullable_percentage,
            "distinct_count": col.distinct_count,
            "duplicate_count": col.duplicate_count,
            "min": col.min,
            "max": col.max,
            "mean": col.mean,
            "median": col.median,
            "sample_values": col.sample_values,
            "description": col.description or "",
            "notes": col.notes or "",
            "tags": col.tags or [],
            "custom_attributes": col.custom_attributes or {}
        })
        
    db_asset = asset_repo.get_by_id(created_asset.id)
    await manager.broadcast({"event_type": "asset_created", "data": {"id": db_asset.id}})
    return db_asset

@router.get("", response_model=List[AssetResponse])
def list_assets(db: Session = Depends(get_db)):
    """
    Retrieves all metadata assets (CSV files).
    """
    return AssetRepository(db).get_all()


def col_letter_to_index(col_letter: str) -> int:
    col_letter = col_letter.upper()
    idx = 0
    for char in col_letter:
        idx = idx * 26 + (ord(char) - ord('A') + 1)
    return idx


@router.post("/profile-preview")
async def profile_preview(files: List[UploadFile] = File(...), db: Session = Depends(get_db)):
    import re
    import uuid
    import io
    from openpyxl import load_workbook
    
    # 1. Fetch all existing columns in the database for fuzzy matching
    existing_cols = db.query(ColumnModel).all()
    
    # Check similarity function (using rapidfuzz if possible, fallback to difflib)
    try:
        from rapidfuzz import fuzz
        def calc_similarity(s1: str, s2: str) -> float:
            return float(fuzz.ratio(s1.lower(), s2.lower()))
    except ImportError:
        import difflib
        def calc_similarity(s1: str, s2: str) -> float:
            return float(difflib.SequenceMatcher(None, s1.lower(), s2.lower()).ratio() * 100.0)
            
    allowed_extensions = (".xlsx", ".xls", ".xlsm")
    proposed_assets = []
    proposed_relationships = []
    
    # A mapping of sheet_name -> list of header names in this upload (to resolve cross-sheet formulas)
    all_sheets_headers = {}
    
    # We will read files and first extract the sheet names and header lists
    file_bytes_map = {}
    for file in files:
        if not file.filename.lower().endswith(allowed_extensions):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"File '{file.filename}' is not a supported Excel workbook format."
            )
        f_bytes = await file.read()
        file_bytes_map[file.filename] = f_bytes
        
        try:
            wb = load_workbook(io.BytesIO(f_bytes), read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Get the first row (headers)
                headers = []
                for cell in ws[1]:
                    if cell.value is not None:
                        headers.append(str(cell.value))
                asset_full_name = f"{file.filename} [{sheet_name}]"
                all_sheets_headers[asset_full_name] = headers
                # Also index by simple sheet name
                all_sheets_headers[sheet_name] = headers
        except Exception as e:
            # Fallback if openpyxl fails to load
            pass

    # Reset file seek positions or just use in-memory map
    for file_name, f_bytes in file_bytes_map.items():
        try:
            # Use openpyxl with data_only=False to inspect formulas
            wb_formula = load_workbook(io.BytesIO(f_bytes), data_only=False)
            # Use profiler to get standard profiling statistics
            profiled_sheets = profile_file(f_bytes, file_name)
            
            # Map sheet name to parsed profile
            # profiled_sheets is a list of (asset_metadata, columns_metadata)
            for asset_meta, cols_meta in profiled_sheets:
                # Find sheet name from asset name e.g. "filename.xlsx [SheetName]"
                sheet_name = asset_meta["name"].split(" [")[-1][:-1] if " [" in asset_meta["name"] else asset_meta["name"]
                
                # Generate a temporary ID for the asset
                temp_asset_id = f"temp_asset_{uuid.uuid4().hex[:8]}"
                asset_meta["temp_id"] = temp_asset_id
                
                # Retrieve the openpyxl worksheet
                ws_f = wb_formula[sheet_name] if sheet_name in wb_formula.sheetnames else None
                
                # Get headers in order
                headers = [c["name"] for c in cols_meta]
                
                # Process columns
                for col_idx, col in enumerate(cols_meta):
                    temp_col_id = f"temp_column_{uuid.uuid4().hex[:8]}"
                    col["temp_id"] = temp_col_id
                    
                    # 2. Check for formula in the first data row (row 2)
                    formula_str = None
                    if ws_f:
                        # columns in openpyxl are 1-indexed
                        cell_val = ws_f.cell(row=2, column=col_idx + 1).value
                        if isinstance(cell_val, str) and cell_val.startswith("="):
                            formula_str = cell_val
                            
                    if formula_str:
                        # Reconstruct formula with column names
                        def replace_ref(match):
                            sheet_q, sheet_n, col_let, row_num = match.groups()
                            ref_sheet = sheet_q or sheet_n or sheet_name
                            c_idx = col_letter_to_index(col_let)
                            
                            if ref_sheet == sheet_name:
                                if 1 <= c_idx <= len(headers):
                                    return f"[{headers[c_idx - 1]}]"
                            else:
                                other_headers = all_sheets_headers.get(ref_sheet)
                                if other_headers and 1 <= c_idx <= len(other_headers):
                                    return f"[{ref_sheet}.{other_headers[c_idx - 1]}]"
                            return match.group(0)
                            
                        pattern = r"(?:(?:'([^']+)'|([A-Za-z0-9_]+))!)?([A-Za-z]+)([0-9]+)"
                        readable_formula = re.sub(pattern, replace_ref, formula_str)
                        
                        col["custom_attributes"]["formula"] = readable_formula
                        
                        # Find unique referenced column indices in same/other sheets to create DERIVES_FROM lineages
                        matches = re.findall(pattern, formula_str)
                        seen_sources = set()
                        for sheet_q, sheet_n, col_let, row_num in matches:
                            ref_sheet = sheet_q or sheet_n or sheet_name
                            c_idx = col_letter_to_index(col_let)
                            
                            source_key = (ref_sheet, c_idx)
                            if source_key in seen_sources:
                                continue
                            seen_sources.add(source_key)
                            
                            # If it's same sheet
                            if ref_sheet == sheet_name:
                                if 1 <= c_idx <= len(headers):
                                    source_col_meta = cols_meta[c_idx - 1]
                                    proposed_relationships.append({
                                        "source_node_type": "column",
                                        "source_node_id": source_col_meta.get("temp_id") or f"col_{c_idx}",
                                        "destination_node_type": "column",
                                        "destination_node_id": temp_col_id,
                                        "relationship_type": "DERIVES_FROM",
                                        "metadata_json": {"formula": readable_formula}
                                    })
                                    
                    # 3. Fuzzy Name Matching against existing database columns
                    # We check all existing columns. If similarity is >= 80%, propose COPIED_FROM
                    for exist_col in existing_cols:
                        score = calc_similarity(col["name"], exist_col.name)
                        if score >= 80.0:
                            proposed_relationships.append({
                                "source_node_type": "column",
                                "source_node_id": exist_col.id,
                                "destination_node_type": "column",
                                "destination_node_id": temp_col_id,
                                "relationship_type": "COPIED_FROM",
                                "metadata_json": {
                                    "similarity": score,
                                    "matched_from": exist_col.name,
                                    "source_table": exist_col.asset.name if exist_col.asset else "Unknown"
                                }
                            })
                            
                asset_meta["columns"] = cols_meta
                proposed_assets.append(asset_meta)
                
            # Now resolve cross-sheet formula lineages since all columns have temp_ids
            # We map (sheet_name, column_index) -> temp_col_id
            col_id_map = {}
            for asset in proposed_assets:
                sheet = asset["name"].split(" [")[-1][:-1] if " [" in asset["name"] else asset["name"]
                for c_idx, col in enumerate(asset["columns"]):
                    col_id_map[(sheet, c_idx + 1)] = col["temp_id"]
                    
            # Let's run a second pass to add cross-sheet formula relationships
            for asset in proposed_assets:
                sheet = asset["name"].split(" [")[-1][:-1] if " [" in asset["name"] else asset["name"]
                ws_f = wb_formula[sheet] if sheet in wb_formula.sheetnames else None
                if not ws_f:
                    continue
                for c_idx, col in enumerate(asset["columns"]):
                    cell_val = ws_f.cell(row=2, column=c_idx + 1).value
                    if isinstance(cell_val, str) and cell_val.startswith("="):
                        formula_str = cell_val
                        pattern = r"(?:(?:'([^']+)'|([A-Za-z0-9_]+))!)?([A-Za-z]+)([0-9]+)"
                        matches = re.findall(pattern, formula_str)
                        seen_sources = set()
                        for sheet_q, sheet_n, col_let, row_num in matches:
                            ref_sheet = sheet_q or sheet_n or sheet
                            src_idx = col_letter_to_index(col_let)
                            
                            # Only handle cross-sheet reference
                            if ref_sheet != sheet:
                                source_key = (ref_sheet, src_idx)
                                if source_key in seen_sources:
                                    continue
                                seen_sources.add(source_key)
                                
                                # Look up the temp column ID
                                src_temp_id = col_id_map.get(source_key)
                                if src_temp_id:
                                    proposed_relationships.append({
                                        "source_node_type": "column",
                                        "source_node_id": src_temp_id,
                                        "destination_node_type": "column",
                                        "destination_node_id": col["temp_id"],
                                        "relationship_type": "DERIVES_FROM",
                                        "metadata_json": {"formula": col["custom_attributes"].get("formula")}
                                    })
                                    
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Failed to profile and extract formulas: {str(e)}"
            )
            
    return {
        "assets": proposed_assets,
        "relationships": proposed_relationships
    }


@router.post("/finalize-import")
async def finalize_import(payload: Dict[str, Any], db: Session = Depends(get_db)):
    asset_repo = AssetRepository(db)
    column_repo = ColumnRepository(db)
    relationship_repo = RelationshipRepository(db)
    
    # We will map temp_id -> actual_db_uuid
    id_map = {}
    created_assets = []
    
    try:
        # 1. Create Assets and Columns
        for asset_data in payload.get("assets", []):
            temp_asset_id = asset_data.get("temp_id")
            
            # Create asset in db
            asset = asset_repo.create(
                name=asset_data["name"],
                asset_type=asset_data.get("asset_type", "excel"),
                row_count=asset_data.get("row_count"),
                column_count=asset_data.get("column_count"),
                file_size=asset_data.get("file_size"),
                description=asset_data.get("description", ""),
                owner=asset_data.get("owner", "Workspace User"),
                notes=asset_data.get("notes", ""),
                tags=asset_data.get("tags", []),
                custom_attributes=asset_data.get("custom_attributes", {})
            )
            
            id_map[temp_asset_id] = asset.id
            
            # Create columns in db
            for col_data in asset_data.get("columns", []):
                temp_col_id = col_data.get("temp_id")
                
                col = column_repo.create(asset.id, {
                    "name": col_data["name"],
                    "datatype": col_data["datatype"],
                    "nullable_percentage": col_data.get("nullable_percentage"),
                    "distinct_count": col_data.get("distinct_count"),
                    "duplicate_count": col_data.get("duplicate_count"),
                    "min": col_data.get("min"),
                    "max": col_data.get("max"),
                    "mean": col_data.get("mean"),
                    "median": col_data.get("median"),
                    "sample_values": col_data.get("sample_values", []),
                    "description": col_data.get("description", ""),
                    "notes": col_data.get("notes", ""),
                    "tags": col_data.get("tags", []),
                    "custom_attributes": col_data.get("custom_attributes", {})
                })
                
                id_map[temp_col_id] = col.id
                
            db_asset = asset_repo.get_by_id(asset.id)
            created_assets.append(db_asset)
            
        # 2. Create Finalized Relationships
        for rel_data in payload.get("relationships", []):
            # Resolve source and target IDs
            raw_source_id = rel_data["source_node_id"]
            raw_dest_id = rel_data["destination_node_id"]
            
            source_id = id_map.get(raw_source_id, raw_source_id)
            dest_id = id_map.get(raw_dest_id, raw_dest_id)
            
            relationship_repo.create(
                source_node_type=rel_data["source_node_type"],
                source_node_id=source_id,
                destination_node_type=rel_data["destination_node_type"],
                destination_node_id=dest_id,
                relationship_type=rel_data["relationship_type"],
                metadata_json=rel_data.get("metadata_json", {})
            )
            
        # Broadcast all asset creations
        for asset in created_assets:
            await manager.broadcast({"event_type": "asset_created", "data": {"id": asset.id}})
            
        # Broadcast relationship sync
        await manager.broadcast({"event_type": "workspace_synced", "data": {}})
        
        return {"status": "success", "message": f"Successfully finalized import of {len(created_assets)} sheets."}
        
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to finalize and save import: {str(e)}"
        )


@router.get("/drafts", response_model=List[ImportDraftResponse])
def list_drafts(db: Session = Depends(get_db)):
    return ImportDraftRepository(db).get_all()


@router.post("/drafts", response_model=ImportDraftResponse)
def save_draft(payload: ImportDraftCreate, db: Session = Depends(get_db)):
    return ImportDraftRepository(db).create(name=payload.name, draft_json=payload.draft_json)


@router.delete("/drafts/{draft_id}")
def delete_draft(draft_id: str, db: Session = Depends(get_db)):
    success = ImportDraftRepository(db).delete(draft_id)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Draft not found"
        )
    return {"status": "success", "message": "Draft successfully deleted."}


@router.get("/{asset_id}", response_model=AssetResponse)
def get_asset(asset_id: str, db: Session = Depends(get_db)):
    """
    Retrieves detailed metadata for a specific asset.
    """
    asset = AssetRepository(db).get_by_id(asset_id)
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Asset with ID '{asset_id}' not found."
        )
    return asset

@router.put("/{asset_id}", response_model=AssetResponse)
async def update_asset(asset_id: str, asset_update: AssetUpdate, db: Session = Depends(get_db)):
    """
    Updates the business metadata (name, description, owner, notes, tags, custom attributes) for an asset.
    Increments the version counter and stores a snapshot in version history.
    """
    asset_repo = AssetRepository(db)
    # Filter out None values to perform partial updates
    updates = asset_update.model_dump(exclude_unset=True)
    
    updated_asset = asset_repo.update(asset_id, updates)
    if not updated_asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Asset with ID '{asset_id}' not found."
        )
    
    # Broadcast asset update event
    await manager.broadcast({"event_type": "asset_updated", "data": {"id": asset_id}})
    
    return updated_asset

@router.delete("/{asset_id}")
async def delete_asset(asset_id: str, db: Session = Depends(get_db)):
    """
    Deletes an asset, its columns, its version history, and all active lineage relationships connected to it.
    """
    success = AssetRepository(db).delete(asset_id)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Asset with ID '{asset_id}' not found."
        )
    
    # Broadcast asset deletion event
    await manager.broadcast({"event_type": "asset_deleted", "data": {"id": asset_id}})
    
    return {"status": "success", "message": f"Asset '{asset_id}' successfully deleted."}

@router.get("/{asset_id}/history", response_model=List[VersionHistoryResponse])
def get_asset_history(asset_id: str, db: Session = Depends(get_db)):
    """
    Retrieves the version history and change logs for a specific asset.
    """
    asset_repo = AssetRepository(db)
    asset = asset_repo.get_by_id(asset_id)
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Asset with ID '{asset_id}' not found."
        )
    return asset_repo.get_version_history(asset_id)

@router.post("/sync")
async def sync_workspace(payload: WorkspaceSync, db: Session = Depends(get_db)):
    """
    Overwrites the current workspace state (assets, columns, relationships) in a single transaction.
    Used for instant session-wide Undo/Redo.
    """
    try:
        # Delete existing data in reverse order of foreign key dependencies
        db.query(RelationshipModel).delete()
        db.query(ColumnModel).delete()
        db.query(Asset).delete()
        db.commit()

        # Insert assets and columns
        for a in payload.assets:
            db_asset = Asset(
                id=a.id,
                name=a.name,
                asset_type=a.asset_type,
                description=a.description or "",
                owner=a.owner or "",
                version=a.version,
                row_count=a.row_count,
                column_count=a.column_count,
                file_size=a.file_size,
                notes=a.notes or "",
                tags=a.tags or [],
                custom_attributes=a.custom_attributes or {},
                created_at=datetime.utcnow(),
                updated_at=datetime.utcnow()
            )
            db.add(db_asset)
            db.commit() # commit so columns are referenceable

            for c in a.columns:
                db_col = ColumnModel(
                    id=c.id,
                    asset_id=db_asset.id,
                    name=c.name,
                    datatype=c.datatype,
                    nullable_percentage=c.nullable_percentage,
                    distinct_count=c.distinct_count,
                    duplicate_count=c.duplicate_count,
                    min=c.min,
                    max=c.max,
                    mean=c.mean,
                    median=c.median,
                    sample_values=c.sample_values or [],
                    description=c.description or "",
                    notes=c.notes or "",
                    tags=c.tags or [],
                    custom_attributes=c.custom_attributes or {},
                    created_at=datetime.utcnow(),
                    updated_at=datetime.utcnow()
                )
                db.add(db_col)
            db.commit()

        # Insert relationships
        for r in payload.relationships:
            db_rel = RelationshipModel(
                id=r.id,
                source_node_type=r.source_node_type,
                source_node_id=r.source_node_id,
                destination_node_type=r.destination_node_type,
                destination_node_id=r.destination_node_id,
                relationship_type=r.relationship_type,
                metadata_json=r.metadata_json or {},
                created_at=datetime.utcnow(),
                updated_at=datetime.utcnow()
            )
            db.add(db_rel)
        db.commit()

        # Broadcast sync event
        await manager.broadcast({"event_type": "workspace_synced", "data": {}})
        return {"status": "success", "message": "Workspace synced successfully."}
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Workspace sync failed: {str(e)}"
        )
